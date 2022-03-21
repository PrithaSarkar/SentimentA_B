from bs4 import BeautifulSoup
import requests
import openpyxl

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',
}

workbook = openpyxl.load_workbook("Input.xlsx")
show = workbook.active

for i in range(2, show.max_row+1):
    url = show.cell(row=i, column=2).value
    print(url)
    url_id = str(int(show.cell(row=i, column=1).value))+".txt"
    print(url_id)
    page = requests.get(url, headers=headers)
    soup = BeautifulSoup(page.content, 'html.parser')
    f = open(url_id, "w", encoding="utf-8")

    for data in soup.find_all(["h1", "p"]):
        sum = data.get_text()
        f.writelines(sum)
    f.close()